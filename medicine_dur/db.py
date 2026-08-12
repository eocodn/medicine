from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import sqlite3
import sys
import time
from contextlib import closing
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .catalog import build_product_catalog
from .product_code_bridge import import_product_code_bridge
from .product_items import import_product_item_sources


PRODUCT_DATASETS = {
    "drug_combination_contraindication.csv": "combination_contraindication",
    "age_contraindication.csv": "age_contraindication",
    "pregnancy_contraindication.csv": "pregnancy_contraindication",
    "dose_caution.csv": "dose_caution",
    "duration_caution.csv": "duration_caution",
    "elderly_caution.csv": "elderly_caution",
    "therapeutic_duplication_caution.csv": "therapeutic_duplication_caution",
}

INGREDIENT_DATASETS = {
    "combination.xlsx": "combination_contraindication",
    "age.xlsx": "age_contraindication",
    "pregnancy.xlsx": "pregnancy_contraindication",
    "dose.xlsx": "dose_caution",
    "duration.xlsx": "duration_caution",
    "elderly.xlsx": "elderly_caution",
    "therapeutic_duplication.xlsx": "therapeutic_duplication_caution",
    "lactation.xlsx": "lactation_caution",
}


SCHEMA = """
PRAGMA journal_mode = DELETE;
PRAGMA synchronous = NORMAL;
PRAGMA temp_store = MEMORY;

CREATE TABLE source_files (
    id INTEGER PRIMARY KEY,
    dataset_key TEXT NOT NULL UNIQUE,
    source_kind TEXT NOT NULL,
    category TEXT NOT NULL,
    source_path TEXT NOT NULL,
    sha256 TEXT NOT NULL,
    size_bytes INTEGER NOT NULL,
    row_count INTEGER NOT NULL,
    imported_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    metadata_json TEXT NOT NULL
);

CREATE TABLE product_dur (
    id INTEGER PRIMARY KEY,
    dataset_key TEXT NOT NULL,
    source_row INTEGER NOT NULL,
    category TEXT NOT NULL,
    ingredient_name TEXT,
    ingredient_code TEXT,
    product_name TEXT,
    product_code TEXT,
    paired_ingredient_name TEXT,
    paired_ingredient_code TEXT,
    paired_product_name TEXT,
    paired_product_code TEXT,
    rule_value TEXT,
    details TEXT,
    notice_no TEXT,
    notice_date TEXT
);

CREATE TABLE product_catalog (
    product_code TEXT PRIMARY KEY,
    product_name TEXT NOT NULL,
    ingredient_code TEXT,
    ingredient_name TEXT
);

CREATE TABLE ingredient_dur (
    id INTEGER PRIMARY KEY,
    dataset_key TEXT NOT NULL,
    source_row INTEGER NOT NULL,
    category TEXT NOT NULL,
    ingredient_name TEXT,
    ingredient_name_ko TEXT,
    paired_ingredient_name TEXT,
    rule_value TEXT,
    dosage_form TEXT,
    note TEXT,
    details TEXT,
    sequence_text TEXT
);

CREATE INDEX idx_product_category ON product_dur(category);
CREATE INDEX idx_product_ingredient_name ON product_dur(ingredient_name);
CREATE INDEX idx_product_ingredient_code ON product_dur(ingredient_code);
CREATE INDEX idx_product_product_code ON product_dur(product_code);
CREATE INDEX idx_product_pair_ingredient_name ON product_dur(paired_ingredient_name);
CREATE INDEX idx_product_pair_ingredient_code ON product_dur(paired_ingredient_code);
CREATE INDEX idx_product_pair_product_code ON product_dur(paired_product_code);
CREATE INDEX idx_catalog_product_name ON product_catalog(product_name);
CREATE INDEX idx_catalog_ingredient_name ON product_catalog(ingredient_name);

CREATE INDEX idx_ingredient_category ON ingredient_dur(category);
CREATE INDEX idx_ingredient_name ON ingredient_dur(ingredient_name);
CREATE INDEX idx_ingredient_name_ko ON ingredient_dur(ingredient_name_ko);
CREATE INDEX idx_ingredient_pair_name ON ingredient_dur(paired_ingredient_name);
"""


def _text(value) -> str | None:
    if value is None:
        return None
    value = str(value).strip()
    return value if value else None


def _code(value) -> str | None:
    value = _text(value)
    if value is None:
        return None
    return re.sub(r"\s+", "", value)


def _therapeutic_duplication_code(value) -> str | None:
    value = _text(value)
    if value is None:
        return None
    # In this specific CSV, embedded blanks are consistently used where a zero
    # appears in the same product/ingredient code in the other DUR datasets.
    return re.sub(r"\s", "0", value)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_record(record: dict) -> str:
    return json.dumps(record, ensure_ascii=False, separators=(",", ":"), default=str)


def _progress(source: str, rows: int, started: float, enabled: bool) -> None:
    if not enabled:
        return
    elapsed = max(time.monotonic() - started, 0.001)
    rate = int(rows / elapsed)
    print(f"[import] {source}: {rows:,} rows ({rate:,}/s)", file=sys.stderr, flush=True)


def _product_common(category: str, row: dict[str, str]) -> tuple:
    ingredient_name = ingredient_code = product_name = product_code = None
    pair_name = pair_code = pair_product_name = pair_product_code = None
    rule_value = details = notice_no = notice_date = None

    if category == "combination_contraindication":
        ingredient_name = _text(row.get("성분명A"))
        ingredient_code = _code(row.get("성분코드A"))
        product_code = _code(row.get("제품코드A"))
        product_name = _text(row.get("제품명A"))
        pair_name = _text(row.get("성분명B"))
        pair_code = _code(row.get("성분코드B"))
        pair_product_code = _code(row.get("제품코드B"))
        pair_product_name = _text(row.get("제품명B"))
        details = _text(row.get("금기사유")) or _text(row.get("비고"))
        notice_no = _text(row.get("고시번호"))
        notice_date = _text(row.get("고시적용일"))
    else:
        ingredient_name = _text(row.get("성분명"))
        ingredient_code = _code(row.get("성분코드"))
        product_code = _code(row.get("제품코드"))
        product_name = _text(row.get("제품명"))
        notice_no = _text(row.get("공고번호"))
        notice_date = _text(row.get("공고일자"))

        if category == "age_contraindication":
            age = _text(row.get("특정연령"))
            unit = _text(row.get("특정연령단위"))
            rule_value = " ".join(v for v in (age, unit) if v) or None
            details = _text(row.get("상세정보"))
        elif category == "pregnancy_contraindication":
            rule_value = _text(row.get("금기등급"))
            details = _text(row.get("상세정보"))
        elif category == "dose_caution":
            rule_value = _text(row.get("1일최대투여량"))
            extra = {
                "1일최대 투여기준량": _text(row.get("1일최대 투여기준량")),
                "점검기준 성분함량 (총함량)": _text(row.get("점검기준 성분함량 (총함량)")),
            }
            details = _json_record({k: v for k, v in extra.items() if v}) if any(extra.values()) else None
        elif category == "duration_caution":
            rule_value = _text(row.get("최대투여기간일수"))
        elif category == "elderly_caution":
            details = _text(row.get("약품상세정보"))
        elif category == "therapeutic_duplication_caution":
            # The published CSV's values are shifted relative to these two headers:
            # 성분코드 contains the ingredient name, while 성분명 contains the ingredient code.
            ingredient_name = _text(row.get("성분코드"))
            ingredient_code = _therapeutic_duplication_code(row.get("성분명"))
            product_code = _therapeutic_duplication_code(row.get("제품코드"))
            rule_value = _text(row.get("효능군"))
            extra = {
                "그룹구분": _text(row.get("그룹구분")),
                "일반명코드": _text(row.get("일반명코드")),
            }
            details = _json_record({k: v for k, v in extra.items() if v}) if any(extra.values()) else None

    return (
        category,
        ingredient_name,
        ingredient_code,
        product_name,
        product_code,
        pair_name,
        pair_code,
        pair_product_name,
        pair_product_code,
        rule_value,
        details,
        notice_no,
        notice_date,
    )


def _import_product_csv(conn: sqlite3.Connection, path: Path, category: str, progress: bool) -> int:
    dataset_key = f"product:{category}"
    started = time.monotonic()
    count = 0
    batch = []
    header = []
    sql = """
        INSERT INTO product_dur (
            dataset_key, source_row, category, ingredient_name, ingredient_code, product_name, product_code,
            paired_ingredient_name, paired_ingredient_code, paired_product_name, paired_product_code,
            rule_value, details, notice_no, notice_date
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """

    with path.open("r", encoding="cp949", newline="") as fh:
        reader = csv.DictReader(fh)
        header = reader.fieldnames or []
        for row in reader:
            count += 1
            common = _product_common(category, row)
            batch.append((dataset_key, count + 1, *common))
            if len(batch) >= 5000:
                conn.executemany(sql, batch)
                batch.clear()
            if count % 100000 == 0:
                _progress(path.name, count, started, progress)
        if batch:
            conn.executemany(sql, batch)

    _progress(path.name, count, started, progress)
    _insert_source_file(conn, dataset_key, "product", category, path, count, {"header": header, "encoding": "cp949"})
    return count


def _find_header(rows: list[tuple], category: str) -> int:
    required = {
        "combination_contraindication": {"연번", "유효성분 '1'", "유효성분 '2'"},
        "age_contraindication": {"연번", "성분명", "연령기준"},
        "pregnancy_contraindication": {"연번", "성분명", "임부금기(등급)"},
        "dose_caution": {"연번", "성분명(국문)", "성분명(영문)", "1일 최대용량"},
        "duration_caution": {"연번", "성분명(국문)", "성분명(영문)", "최대 투여기간"},
        "elderly_caution": {"연번", "성분명(국문)", "성분명(영문)"},
        "therapeutic_duplication_caution": {"연번", "효능군", "성분명(국문)", "성분명(영문)"},
        "lactation_caution": {"연번", "성분명(국문)", "성분명(영문)"},
    }[category]
    for idx, row in enumerate(rows):
        values = {_text(v) for v in row if _text(v)}
        if required.issubset(values):
            return idx
    raise ValueError(f"Could not find header row for {category}")


def _ingredient_common(category: str, row: dict[str, object]) -> tuple:
    ingredient_name = ingredient_name_ko = pair_name = rule_value = dosage_form = note = details = sequence = None

    if category == "combination_contraindication":
        sequence = _text(row.get("연번"))
        ingredient_name = _text(row.get("유효성분 '1'"))
        pair_name = _text(row.get("유효성분 '2'"))
        note = _text(row.get("비고"))
        details = _text(row.get("상세정보"))
    elif category == "age_contraindication":
        sequence = _text(row.get("연번"))
        ingredient_name = _text(row.get("성분명"))
        rule_value = _text(row.get("연령기준"))
        dosage_form = _text(row.get("제형"))
        details = _text(row.get("상세정보"))
    elif category == "pregnancy_contraindication":
        sequence = _text(row.get("연번"))
        ingredient_name = _text(row.get("성분명"))
        rule_value = _text(row.get("임부금기(등급)"))
        note = _text(row.get("비고"))
        details = _text(row.get("상세정보"))
    else:
        sequence = _text(row.get("연번"))
        ingredient_name_ko = _text(row.get("성분명(국문)"))
        ingredient_name = _text(row.get("성분명(영문)"))
        dosage_form = _text(row.get("제형"))
        note = _text(row.get("비고"))

        if category == "dose_caution":
            rule_value = _text(row.get("1일 최대용량"))
        elif category == "duration_caution":
            rule_value = _text(row.get("최대 투여기간"))
        elif category == "therapeutic_duplication_caution":
            sequence = _text(row.get("연번")) or _text(row.get("연번_3"))
            rule_value = _text(row.get("효능군"))
            series = _text(row.get("계열명"))
            if series:
                note = f"{series}\n{note}" if note else series

    return (
        category,
        ingredient_name,
        ingredient_name_ko,
        pair_name,
        rule_value,
        dosage_form,
        note,
        details,
        sequence,
    )


def _dedupe_headers(values: Iterable[object]) -> list[str]:
    counts: dict[str, int] = {}
    result = []
    for index, value in enumerate(values):
        base = _text(value) or f"column_{index + 1}"
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return result


def _import_ingredient_xlsx(conn: sqlite3.Connection, path: Path, category: str, progress: bool) -> int:
    dataset_key = f"ingredient:{category}"
    started = time.monotonic()
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    header_index = _find_header(rows[:10], category)
    headers = _dedupe_headers(rows[header_index])
    count = 0
    batch = []
    sql = """
        INSERT INTO ingredient_dur (
            dataset_key, source_row, category, ingredient_name, ingredient_name_ko, paired_ingredient_name,
            rule_value, dosage_form, note, details, sequence_text
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """

    for source_row, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not any(_text(v) for v in values):
            continue
        record = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        common = _ingredient_common(category, record)
        if not common[1] and not common[2]:
            continue
        count += 1
        batch.append((dataset_key, source_row, *common))
        if len(batch) >= 1000:
            conn.executemany(sql, batch)
            batch.clear()
    if batch:
        conn.executemany(sql, batch)
    workbook.close()

    _progress(path.name, count, started, progress)
    metadata = {
        "sheet": worksheet.title,
        "header_row": header_index + 1,
        "header": headers,
        "title": _text(rows[0][0]) if rows and rows[0] else None,
    }
    _insert_source_file(conn, dataset_key, "ingredient", category, path, count, metadata)
    return count


def _insert_source_file(
    conn: sqlite3.Connection,
    dataset_key: str,
    source_kind: str,
    category: str,
    path: Path,
    row_count: int,
    metadata: dict,
) -> None:
    conn.execute(
        """
        INSERT INTO source_files (
            dataset_key, source_kind, category, source_path, sha256, size_bytes, row_count, metadata_json
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            dataset_key,
            source_kind,
            category,
            str(path),
            _sha256(path),
            path.stat().st_size,
            row_count,
            _json_record(metadata),
        ),
    )



def build_database(
    db_path: str | Path,
    raw_dir: str | Path,
    kids_dir: str | Path,
    *,
    progress: bool = True,
) -> dict:
    db_path = Path(db_path)
    raw_dir = Path(raw_dir)
    kids_dir = Path(kids_dir)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = db_path.with_name(db_path.name + ".tmp")
    if temp_path.exists():
        temp_path.unlink()

    product_rows = 0
    ingredient_rows = 0
    product_flag_rows = 0
    product_bridge_rows = 0
    catalog_products = 0
    source_files = 0
    started = time.monotonic()

    try:
        with closing(sqlite3.connect(temp_path)) as conn:
            conn.executescript(SCHEMA)
            conn.execute("BEGIN")

            for filename, category in PRODUCT_DATASETS.items():
                path = raw_dir / filename
                if not path.exists():
                    continue
                product_rows += _import_product_csv(conn, path, category, progress)
                source_files += 1
                conn.commit()
                conn.execute("BEGIN")

            for filename, category in INGREDIENT_DATASETS.items():
                path = kids_dir / filename
                if not path.exists():
                    continue
                ingredient_rows += _import_ingredient_xlsx(conn, path, category, progress)
                source_files += 1
                conn.commit()
                conn.execute("BEGIN")

            item_sources, product_flag_rows = import_product_item_sources(conn, raw_dir)
            source_files += item_sources

            bridge_sources, product_bridge_rows = import_product_code_bridge(conn, raw_dir)
            source_files += bridge_sources

            catalog_products = build_product_catalog(conn)
            conn.commit()
            conn.execute("ANALYZE")
            conn.execute("PRAGMA optimize")

        if source_files == 0:
            raise ValueError("No recognized DUR source files found")
        os.replace(temp_path, db_path)
    except Exception:
        if temp_path.exists():
            temp_path.unlink()
        raise

    return {
        "db_path": str(db_path),
        "source_files": source_files,
        "product_rows": product_rows,
        "ingredient_rows": ingredient_rows,
        "product_flag_rows": product_flag_rows,
        "product_bridge_rows": product_bridge_rows,
        "catalog_products": catalog_products,
        "total_rows": product_rows + ingredient_rows + product_flag_rows + product_bridge_rows,
        "elapsed_seconds": round(time.monotonic() - started, 3),
        "size_bytes": db_path.stat().st_size,
    }


def database_stats(db_path: str | Path) -> dict:
    db_path = Path(db_path)
    with closing(sqlite3.connect(db_path)) as conn:
        conn.row_factory = sqlite3.Row
        product_rows = conn.execute("SELECT COUNT(*) AS n FROM product_dur").fetchone()["n"]
        ingredient_rows = conn.execute("SELECT COUNT(*) AS n FROM ingredient_dur").fetchone()["n"]
        product_flag_rows = (
            conn.execute("SELECT COUNT(*) AS n FROM product_item_flags").fetchone()["n"]
            if conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='product_item_flags'"
            ).fetchone()
            else 0
        )
        product_bridge_rows = (
            conn.execute("SELECT COUNT(*) AS n FROM product_code_bridge").fetchone()["n"]
            if conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='product_code_bridge'"
            ).fetchone()
            else 0
        )
        catalog_products = conn.execute("SELECT COUNT(*) AS n FROM product_catalog").fetchone()["n"]
        source_files = conn.execute("SELECT COUNT(*) AS n FROM source_files").fetchone()["n"]
        categories = [
            dict(row)
            for row in conn.execute(
                """
                SELECT category, source_kind, SUM(row_count) AS rows
                FROM source_files
                GROUP BY category, source_kind
                ORDER BY category, source_kind
                """
            )
        ]
    return {
        "db_path": str(db_path),
        "source_files": source_files,
        "product_rows": product_rows,
        "ingredient_rows": ingredient_rows,
        "product_flag_rows": product_flag_rows,
        "product_bridge_rows": product_bridge_rows,
        "catalog_products": catalog_products,
        "total_rows": product_rows + ingredient_rows + product_flag_rows + product_bridge_rows,
        "size_bytes": db_path.stat().st_size,
        "categories": categories,
    }


def search_records(db_path: str | Path, term: str, *, limit: int = 20) -> list[dict]:
    if not term.strip():
        raise ValueError("Search term must not be empty")
    pattern = f"%{term.strip()}%"
    sql = """
        SELECT
            'product' AS source_kind,
            category,
            ingredient_name,
            NULL AS ingredient_name_ko,
            paired_ingredient_name,
            product_name,
            paired_product_name,
            rule_value,
            details
        FROM product_dur
        WHERE ingredient_name LIKE ? OR paired_ingredient_name LIKE ?
           OR product_name LIKE ? OR paired_product_name LIKE ?
        UNION ALL
        SELECT
            'ingredient' AS source_kind,
            category,
            ingredient_name,
            ingredient_name_ko,
            paired_ingredient_name,
            NULL AS product_name,
            NULL AS paired_product_name,
            rule_value,
            details
        FROM ingredient_dur
        WHERE ingredient_name LIKE ? OR ingredient_name_ko LIKE ? OR paired_ingredient_name LIKE ?
        LIMIT ?
    """
    with closing(sqlite3.connect(db_path)) as conn:
        conn.row_factory = sqlite3.Row
        params = [pattern, pattern, pattern, pattern, pattern, pattern, pattern, limit]
        return [dict(row) for row in conn.execute(sql, params)]
