from __future__ import annotations

import hashlib
import json
import re
import sqlite3
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

XLSX_DATASETS: dict[str, str] = {
    "combination.xlsx": "combination_contraindication",
    "age.xlsx": "age_contraindication",
    "pregnancy.xlsx": "pregnancy_contraindication",
    "dose.xlsx": "dose_caution",
    "duration.xlsx": "duration_caution",
    "elderly.xlsx": "elderly_caution",
    "therapeutic_duplication.xlsx": "therapeutic_duplication_caution",
    "lactation.xlsx": "lactation_caution",
}


def _text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json(record: dict) -> str:
    return json.dumps(record, ensure_ascii=False, separators=(",", ":"), sort_keys=True, default=str)


def _dedupe_headers(values: Iterable[object]) -> list[str]:
    counts: dict[str, int] = {}
    result: list[str] = []
    for index, value in enumerate(values):
        base = _text(value) or f"column_{index + 1}"
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return result


def _find_header(rows: list[tuple], category: str) -> int:
    required = {
        "combination_contraindication": {"연번", "유효성분 '1'", "유효성분 '2'"},
        "age_contraindication": {"연번", "성분명", "연령기준"},
        "pregnancy_contraindication": {"연번", "성분명", "임부금기(등급)"},
        "dose_caution": {"연번", "성분명(국문)", "성분명(영문)", "1일 최대용량"},
        "duration_caution": {"연번", "성분명(국문)", "성분명(영문)", "최대 투여기간"},
        "elderly_caution": {"연번", "성분명(국문)", "성분명(영문)"},
        "therapeutic_duplication_caution": {"연번", "효능군", "성분명(국문)", "성분명(영문)"},
        "lactation_caution": {"연번", "성분명(국문)", "성분명(영문)"},
    }[category]
    for idx, row in enumerate(rows):
        values = {_text(v) for v in row if _text(v)}
        if required.issubset(values):
            return idx
    raise ValueError(f"could not find XLSX header for {category}")


def _effective_date(title: str | None) -> str | None:
    if not title:
        return None
    match = re.search(r"(20\d{2})[.\-/](\d{1,2})[.\-/](\d{1,2})", title)
    if not match:
        return None
    year, month, day = (int(part) for part in match.groups())
    return f"{year:04d}-{month:02d}-{day:02d}"


def inspect_xlsx_source(path: str | Path, category: str) -> dict:
    source = Path(path)
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True, max_row=10))
        header_idx = _find_header(rows, category)
        headers = _dedupe_headers(rows[header_idx])
        title = _text(rows[0][0]) if rows and rows[0] else None
        return {
            "sheet": worksheet.title,
            "header_row": header_idx + 1,
            "header": headers,
            "title": title,
            "effective_date": _effective_date(title),
        }
    finally:
        workbook.close()


def _canonical_rule(category: str, row: dict[str, object]) -> dict:
    result = {
        "category": category,
        "sequence_text": None,
        "ingredient_name": None,
        "ingredient_name_ko": None,
        "paired_ingredient_name": None,
        "rule_value": None,
        "dosage_form": None,
        "note": None,
        "details": None,
    }
    if category == "combination_contraindication":
        result.update(
            sequence_text=_text(row.get("연번")),
            ingredient_name=_text(row.get("유효성분 '1'")),
            paired_ingredient_name=_text(row.get("유효성분 '2'")),
            note=_text(row.get("비고")),
            details=_text(row.get("상세정보")),
        )
    elif category == "age_contraindication":
        result.update(
            sequence_text=_text(row.get("연번")),
            ingredient_name=_text(row.get("성분명")),
            rule_value=_text(row.get("연령기준")),
            dosage_form=_text(row.get("제형")),
            details=_text(row.get("상세정보")),
        )
    elif category == "pregnancy_contraindication":
        result.update(
            sequence_text=_text(row.get("연번")),
            ingredient_name=_text(row.get("성분명")),
            rule_value=_text(row.get("임부금기(등급)")),
            note=_text(row.get("비고")),
            details=_text(row.get("상세정보")),
        )
    else:
        result.update(
            sequence_text=_text(row.get("연번")),
            ingredient_name=_text(row.get("성분명(영문)")),
            ingredient_name_ko=_text(row.get("성분명(국문)")),
            dosage_form=_text(row.get("제형")),
            note=_text(row.get("비고")),
        )
        if category == "dose_caution":
            result["rule_value"] = _text(row.get("1일 최대용량"))
        elif category == "duration_caution":
            result["rule_value"] = _text(row.get("최대 투여기간"))
        elif category == "therapeutic_duplication_caution":
            result["sequence_text"] = _text(row.get("연번_3")) or _text(row.get("연번"))
            result["rule_value"] = _text(row.get("효능군"))
            series = _text(row.get("계열명"))
            if series:
                result["note"] = f"{series}\n{result['note']}" if result["note"] else series
    return result


def import_xlsx_sources(con: sqlite3.Connection, kids_dir: str | Path) -> dict:
    root = Path(kids_dir)
    missing = [filename for filename in XLSX_DATASETS if not (root / filename).exists()]
    if missing:
        raise FileNotFoundError(f"missing required KIDS/MFDS XLSX sources: {', '.join(sorted(missing))}")

    total_rows = 0
    sources: list[dict] = []
    for filename, category in XLSX_DATASETS.items():
        path = root / filename
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
            header_idx = _find_header(rows[:10], category)
            headers = _dedupe_headers(rows[header_idx])
            title = _text(rows[0][0]) if rows and rows[0] else None
            dataset_key = f"kids_mfds_xlsx:{category}"
            con.execute(
                """
                INSERT INTO source_snapshots(
                    dataset_key,source_family,source_locator,snapshot_path,title,effective_date,row_count,
                    reported_row_count,sha256,metadata_json
                ) VALUES(?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    dataset_key,
                    "kids_mfds_xlsx",
                    str(path),
                    str(path),
                    title,
                    _effective_date(title),
                    0,
                    0,
                    _sha256(path),
                    "{}",
                ),
            )
            imported = 0
            for source_row, values in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
                if not any(_text(v) for v in values):
                    continue
                record = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
                canonical = _canonical_rule(category, record)
                if not canonical["ingredient_name"] and not canonical["ingredient_name_ko"]:
                    continue
                con.execute(
                    """
                    INSERT INTO ingredient_rules(
                        source_dataset_key,source_row,category,sequence_text,ingredient_name,
                        ingredient_name_ko,paired_ingredient_name,rule_value,dosage_form,note,details
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        dataset_key,
                        source_row,
                        canonical["category"],
                        canonical["sequence_text"],
                        canonical["ingredient_name"],
                        canonical["ingredient_name_ko"],
                        canonical["paired_ingredient_name"],
                        canonical["rule_value"],
                        canonical["dosage_form"],
                        canonical["note"],
                        canonical["details"],
                    ),
                )
                imported += 1
            metadata = {
                "sheet": worksheet.title,
                "header_row": header_idx + 1,
                "header": headers,
                "title": title,
            }
            con.execute(
                "UPDATE source_snapshots SET row_count=?, reported_row_count=?, metadata_json=? WHERE dataset_key=?",
                (imported, imported, _json(metadata), dataset_key),
            )
            sources.append({"dataset_key": dataset_key, "category": category, "rows": imported, "effective_date": _effective_date(title)})
            total_rows += imported
        finally:
            workbook.close()
    return {"ingredient_rules": total_rows, "sources": sources}
